import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook


file_path = r"C:\Users\anshu\Desktop\Anshul\University\University of Queensland\REIT4842 - R&D Methods and Practice\Proposal Draft\pypi100.xlsx"

try:
    workbook = load_workbook(filename=file_path)
    sheet_names = workbook.sheetnames
    print("Sheet names:", sheet_names)
    
    sheet = workbook[sheet_names[0]]
    
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    
    
    columns = data[0]  
    df = pd.DataFrame(data[1:], columns=columns)
    df.columns = [str(col).strip() for col in df.columns] 
    print("Columns in DataFrame:", df.columns)
    print(df.head())
    
    repo_correctness = df['Is the repo link correct?'].value_counts()

    plt.figure(figsize=(10, 6))
    repo_correctness.plot(kind='bar', color=['#67ede7', '#fce85d', '#ff9933']) 
    plt.title('Correct vs Incorrect Repository Links (PyPI)')
    plt.xlabel('Repository Link Correctness')
    plt.ylabel('Number of Packages')
    plt.xticks(rotation=0)
    plt.grid(axis='y', alpha=0.3) 
    plt.show()

    related_package_repo = df['Does it have the repo as related package?'].value_counts()
    plt.figure(figsize=(8, 8))
    related_package_repo.plot(kind='pie', autopct='%1.1f%%', startangle=140, colors=['#aaf42c', '#52f24a'])
    plt.title('Packages declaring the main repository in their metadata (PyPI)')
    plt.ylabel('')
    plt.show()

    registry_link_correctness = df['Is it correct?'].value_counts()
    plt.figure(figsize=(8, 8))
    registry_link_correctness.plot(kind='pie', autopct='%1.1f%%', startangle=140, colors=['#ff3377', '#4c00a4'])  
    plt.title('Correctness of PyPI package registry link collected by deps.dev')
    plt.ylabel('')
    plt.show()

except Exception as e:
    print(f"Error: {e}")
